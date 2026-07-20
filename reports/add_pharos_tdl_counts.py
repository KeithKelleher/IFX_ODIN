from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import pymysql


WORKBOOK = "reports/pharos_tdl_changes_by_uniprot_since_2017.xlsx"
SCHEMAS = [
    "tcrd322",
    "tcrd540",
    "tcrd660",
    "tcrd6110",
    "pharos316",
    "pharos317",
    "pharos318",
    "pharos319",
    "pharos400",
]
TDLS = ["Tclin", "Tchem", "Tbio", "Tdark"]
PAIRS = [
    ("tcrd322", "tcrd540"),
    ("tcrd540", "tcrd660"),
    ("tcrd660", "tcrd6110"),
    ("tcrd6110", "pharos316"),
    ("pharos316", "pharos317"),
    ("pharos317", "pharos318"),
    ("pharos318", "pharos319"),
    ("pharos319", "pharos400"),
]


def main():
    conn = pymysql.connect(host="tcrd.ncats.io", user="tcrd")
    cur = conn.cursor()
    rows = []
    schema_maps = {}
    for schema in SCHEMAS:
        cur.execute(
            f"""
            SELECT
              p.uniprot,
              CASE t.tdl
                WHEN 'TclinPlus' THEN 'Tclin'
                WHEN 'TchemPlus' THEN 'Tchem'
                ELSE t.tdl
              END AS tdl
            FROM {schema}.target t
            JOIN {schema}.t2tc c ON c.target_id = t.id
            JOIN {schema}.protein p ON p.id = c.protein_id
            WHERE p.uniprot IS NOT NULL
              AND p.uniprot <> ''
              AND t.tdl IS NOT NULL
            """
        )
        schema_maps[schema] = {}
        for uniprot, tdl in cur.fetchall():
            schema_maps[schema][uniprot] = tdl
        counts = {}
        for tdl in schema_maps[schema].values():
            counts[tdl] = counts.get(tdl, 0) + 1
        for tdl in TDLS:
            rows.append((schema, tdl, counts.get(tdl, 0)))

    change_set_rows = []
    for old_schema, new_schema in PAIRS:
        old_map = schema_maps[old_schema]
        new_map = schema_maps[new_schema]
        removed = {}
        for uniprot, tdl in old_map.items():
            if uniprot not in new_map:
                removed[tdl] = removed.get(tdl, 0) + 1
        for tdl in TDLS:
            if removed.get(tdl, 0):
                change_set_rows.append((f"{old_schema}->{new_schema}", "removed", tdl, removed[tdl]))

        added = {}
        for uniprot, tdl in new_map.items():
            if uniprot not in old_map:
                added[tdl] = added.get(tdl, 0) + 1
        for tdl in TDLS:
            if added.get(tdl, 0):
                change_set_rows.append((f"{old_schema}->{new_schema}", "added", tdl, added[tdl]))
    conn.close()

    wb = load_workbook(WORKBOOK)
    if "tdl_counts" in wb.sheetnames:
        del wb["tdl_counts"]
    idx = wb.sheetnames.index("transition_summary")
    ws = wb.create_sheet("tdl_counts", idx)
    ws.append(["schema", "tdl", "uniprot_count"])
    for row in rows:
        ws.append(row)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:C{ws.max_row}"
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 16

    if "target_set_changes" in wb.sheetnames:
        del wb["target_set_changes"]
    idx = wb.sheetnames.index("transition_summary")
    ws = wb.create_sheet("target_set_changes", idx)
    ws.append(["upgrade", "change_type", "tdl", "uniprot_count"])
    for row in change_set_rows:
        ws.append(row)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="FCE4D6")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:D{ws.max_row}"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 16
    wb.save(WORKBOOK)

    print(WORKBOOK)
    for row in rows:
        print(row)
    for row in change_set_rows:
        print(row)


if __name__ == "__main__":
    main()
