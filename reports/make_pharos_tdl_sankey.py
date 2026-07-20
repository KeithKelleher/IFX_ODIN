from collections import defaultdict
from html import escape
import json
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path("reports/pharos_tdl_changes_by_uniprot_since_2017.xlsx")
SVG_OUT = Path("reports/pharos_tdl_changes_sankey.svg")
HTML_OUT = Path("reports/pharos_tdl_changes_sankey.html")
PLOTLY_OUT = Path("reports/pharos_tdl_changes_sankey_plotly.html")

TDL_ORDER = ["Tclin", "Tchem", "Tbio", "Tdark", "Tvoid"]
TDL_COLORS = {
    "Tclin": "#d73027",
    "Tchem": "#fc8d59",
    "Tbio": "#91bfdb",
    "Tdark": "#636363",
    "Tvoid": "#d9d9d9",
}
PAIR_SHEETS = {
    ("tcrd322", "tcrd540"): "322_to_540",
    ("tcrd540", "tcrd660"): "540_to_660",
    ("tcrd660", "tcrd6110"): "660_to_6110",
    ("tcrd6110", "pharos316"): "6110_to_316",
    ("pharos316", "pharos317"): "316_to_317",
    ("pharos317", "pharos318"): "317_to_318",
    ("pharos318", "pharos319"): "318_to_319",
    ("pharos319", "pharos400"): "319_to_400",
}


def normalize_tdl(value):
    if value == "TclinPlus":
        return "Tclin"
    if value == "TchemPlus":
        return "Tchem"
    return value


def parse_summary():
    wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
    ws = wb["transition_summary"]
    rows = []
    stages = []
    for upgrade, transition, count in ws.iter_rows(min_row=2, values_only=True):
        if not upgrade or not transition or not count:
            continue
        old_schema, new_schema = upgrade.split("->")
        old_tdl, new_tdl = [part.strip() for part in transition.split("->")]
        if old_tdl == "no TDL changes":
            continue
        old_tdl = normalize_tdl(old_tdl)
        new_tdl = normalize_tdl(new_tdl)
        rows.append((old_schema, new_schema, old_tdl, new_tdl, int(count)))
        if old_schema not in stages:
            stages.append(old_schema)
        if new_schema not in stages:
            stages.append(new_schema)
    return stages, rows


def parse_full_flows():
    wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
    stages = [row[0] for row in wb["metadata_dates"].iter_rows(min_row=2, values_only=True) if row[0]]
    totals = defaultdict(lambda: defaultdict(int))
    for schema, tdl, count in wb["tdl_counts"].iter_rows(min_row=2, values_only=True):
        if schema and tdl:
            totals[schema][normalize_tdl(tdl)] = int(count or 0)

    target_set_changes = defaultdict(lambda: defaultdict(int))
    if "target_set_changes" in wb.sheetnames:
        for upgrade, change_type, tdl, count in wb["target_set_changes"].iter_rows(min_row=2, values_only=True):
            if upgrade and change_type and tdl and count:
                target_set_changes[(upgrade, change_type)][normalize_tdl(tdl)] += int(count)

    rows = []
    for old_schema, new_schema in zip(stages, stages[1:]):
        # Plotly drops/re-packs zero-value nodes. A tiny transparent Tvoid
        # anchor keeps the bottom lane present even when no targets enter/exit.
        rows.append((old_schema, new_schema, "Tvoid", "Tvoid", 0.001))
        sheet = PAIR_SHEETS.get((old_schema, new_schema))
        if sheet not in wb.sheetnames:
            continue
        changed = {}
        for uniprot, symbol, name, old_tdl, new_tdl, transition in wb[sheet].iter_rows(min_row=2, values_only=True):
            if not uniprot or not old_tdl or not new_tdl:
                continue
            old_tdl = normalize_tdl(old_tdl)
            new_tdl = normalize_tdl(new_tdl)
            changed[(old_tdl, new_tdl)] = changed.get((old_tdl, new_tdl), 0) + 1

        upgrade = f"{old_schema}->{new_schema}"
        removed = target_set_changes[(upgrade, "removed")]
        added = target_set_changes[(upgrade, "added")]

        outgoing_changed = defaultdict(int)
        for (old_tdl, new_tdl), count in changed.items():
            rows.append((old_schema, new_schema, old_tdl, new_tdl, count))
            outgoing_changed[old_tdl] += count

        for tdl in [tdl for tdl in TDL_ORDER if tdl != "Tvoid"]:
            stayed = totals[old_schema][tdl] - outgoing_changed[tdl] - removed[tdl]
            if stayed < 0:
                raise RuntimeError(f"Changed count exceeds total for {old_schema} {tdl}")
            if stayed:
                rows.append((old_schema, new_schema, tdl, tdl, stayed))
            if removed[tdl]:
                rows.append((old_schema, new_schema, tdl, "Tvoid", removed[tdl]))
            if added[tdl]:
                rows.append((old_schema, new_schema, "Tvoid", tdl, added[tdl]))
    return stages, rows


def path(x1, y1, x2, y2):
    dx = x2 - x1
    return f"M {x1:.1f},{y1:.1f} C {x1 + dx * 0.55:.1f},{y1:.1f} {x2 - dx * 0.55:.1f},{y2:.1f} {x2:.1f},{y2:.1f}"


def render_svg(stages, rows):
    width = 1680
    height = 860
    margin_x = 90
    top = 125
    lane_gap = 160
    node_w = 18
    node_h = 86
    flow_scale = 0.050
    min_flow_w = 1.2

    x_by_stage = {
        stage: margin_x + i * ((width - 2 * margin_x) / (len(stages) - 1))
        for i, stage in enumerate(stages)
    }
    y_by_tdl = {tdl: top + i * lane_gap for i, tdl in enumerate(TDL_ORDER)}

    out_offsets = defaultdict(float)
    in_offsets = defaultdict(float)
    node_counts = defaultdict(int)
    pair_counts = defaultdict(int)
    for old_schema, new_schema, old_tdl, new_tdl, count in rows:
        node_counts[(old_schema, old_tdl)] += count
        node_counts[(new_schema, new_tdl)] += count
        pair_counts[(old_schema, new_schema)] += count

    # Draw bigger flows first so smaller changes remain visible on top.
    sorted_rows = sorted(rows, key=lambda r: r[4], reverse=True)
    flow_elements = []
    for old_schema, new_schema, old_tdl, new_tdl, count in sorted_rows:
        old_key = (old_schema, old_tdl)
        new_key = (new_schema, new_tdl)
        stroke_w = max(min_flow_w, count * flow_scale)
        x1 = x_by_stage[old_schema] + node_w
        x2 = x_by_stage[new_schema]
        y1 = y_by_tdl[old_tdl] + 18 + out_offsets[old_key] + stroke_w / 2
        y2 = y_by_tdl[new_tdl] + 18 + in_offsets[new_key] + stroke_w / 2
        out_offsets[old_key] += stroke_w + 1.2
        in_offsets[new_key] += stroke_w + 1.2
        color = TDL_COLORS.get(old_tdl, "#888")
        title = f"{old_schema} {old_tdl} -> {new_schema} {new_tdl}: {count:,} UniProt targets"
        flow_elements.append(
            f'<path d="{path(x1, y1, x2, y2)}" fill="none" stroke="{color}" '
            f'stroke-width="{stroke_w:.2f}" stroke-opacity="0.42" '
            f'stroke-linecap="round"><title>{escape(title)}</title></path>'
        )

    node_elements = []
    label_elements = []
    for stage in stages:
        x = x_by_stage[stage]
        label_elements.append(
            f'<text x="{x + node_w / 2:.1f}" y="55" text-anchor="middle" class="stage">{escape(stage)}</text>'
        )
        for tdl in TDL_ORDER:
            y = y_by_tdl[tdl]
            count = node_counts[(stage, tdl)]
            node_elements.append(
                f'<rect x="{x:.1f}" y="{y:.1f}" width="{node_w}" height="{node_h}" rx="3" '
                f'fill="{TDL_COLORS[tdl]}"><title>{escape(stage)} {tdl}: {count:,} changed-flow endpoints</title></rect>'
            )
            label_elements.append(
                f'<text x="{x + node_w + 7:.1f}" y="{y + 21:.1f}" class="tdl">{tdl}</text>'
            )
            if count:
                label_elements.append(
                    f'<text x="{x + node_w + 7:.1f}" y="{y + 41:.1f}" class="count">{count:,}</text>'
                )

    legend_x = 90
    legend_y = 790
    legend = []
    for i, tdl in enumerate(TDL_ORDER):
        x = legend_x + i * 135
        legend.append(f'<rect x="{x}" y="{legend_y}" width="18" height="18" rx="3" fill="{TDL_COLORS[tdl]}"/>')
        legend.append(f'<text x="{x + 26}" y="{legend_y + 14}" class="legend">{tdl}</text>')

    pair_labels = []
    for (old_schema, new_schema), count in pair_counts.items():
        x = (x_by_stage[old_schema] + x_by_stage[new_schema]) / 2
        pair_labels.append(
            f'<text x="{x:.1f}" y="91" text-anchor="middle" class="pair">{count:,} changed</text>'
        )

    css = """
    .title { font: 700 24px Arial, sans-serif; fill: #1f2933; }
    .subtitle { font: 14px Arial, sans-serif; fill: #52616b; }
    .stage { font: 700 13px Arial, sans-serif; fill: #1f2933; }
    .pair { font: 12px Arial, sans-serif; fill: #52616b; }
    .tdl { font: 700 12px Arial, sans-serif; fill: #1f2933; }
    .count { font: 11px Arial, sans-serif; fill: #52616b; }
    .legend { font: 13px Arial, sans-serif; fill: #1f2933; }
    """
    svg = f"""<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">
<style>{css}</style>
<rect width="100%" height="100%" fill="#ffffff"/>
<text x="90" y="38" class="title">TDL Changes by UniProt Across TCRD / Pharos Versions</text>
<text x="90" y="64" class="subtitle">Ribbon width is proportional to changed UniProt target count. Colors indicate the source TDL before each upgrade.</text>
{''.join(pair_labels)}
{''.join(flow_elements)}
{''.join(node_elements)}
{''.join(label_elements)}
{''.join(legend)}
</svg>
"""
    return svg


def rgba(hex_color, alpha):
    hex_color = hex_color.lstrip("#")
    r = int(hex_color[0:2], 16)
    g = int(hex_color[2:4], 16)
    b = int(hex_color[4:6], 16)
    return f"rgba({r},{g},{b},{alpha})"


def render_plotly_html(stages, rows):
    node_values = defaultdict(float)
    outgoing = defaultdict(float)
    incoming = defaultdict(float)
    for old_schema, new_schema, old_tdl, new_tdl, count in rows:
        outgoing[(old_schema, old_tdl)] += count
        incoming[(new_schema, new_tdl)] += count
    for stage in stages:
        for tdl in TDL_ORDER:
            node_values[(stage, tdl)] = max(incoming[(stage, tdl)], outgoing[(stage, tdl)])

    y_by_node = {}
    gap = 0.015
    top = 0.015
    usable = 0.94 - gap * (len(TDL_ORDER) - 1)
    for stage in stages:
        total = sum(node_values[(stage, tdl)] for tdl in TDL_ORDER) or 1
        cursor = top
        for tdl in TDL_ORDER:
            y_by_node[(stage, tdl)] = cursor
            cursor += usable * (node_values[(stage, tdl)] / total) + gap

    node_index = {}
    labels = []
    colors = []
    xs = []
    ys = []
    customdata = []

    for stage_i, stage in enumerate(stages):
        x = 0.02 + stage_i * (0.96 / max(1, len(stages) - 1))
        for tdl_i, tdl in enumerate(TDL_ORDER):
            node_index[(stage, tdl)] = len(labels)
            labels.append(tdl)
            colors.append(TDL_COLORS[tdl])
            xs.append(x)
            ys.append(y_by_node[(stage, tdl)])
            customdata.append(f"{stage} {tdl}")

    sources = []
    targets = []
    values = []
    link_colors = []
    link_customdata = []
    link_hovertemplates = []
    for old_schema, new_schema, old_tdl, new_tdl, count in rows:
        sources.append(node_index[(old_schema, old_tdl)])
        targets.append(node_index[(new_schema, new_tdl)])
        values.append(count)
        is_anchor = old_tdl == "Tvoid" and new_tdl == "Tvoid" and count < 1
        link_colors.append("rgba(0,0,0,0)" if is_anchor else rgba(TDL_COLORS.get(old_tdl, "#888888"), 0.35))
        if old_tdl == "Tvoid":
            transition_label = f"Tvoid -> {new_tdl} newly annotated"
        elif new_tdl == "Tvoid":
            transition_label = f"{old_tdl} -> Tvoid removed/obsolete"
        elif old_tdl == new_tdl:
            transition_label = f"{old_tdl} unchanged"
        else:
            transition_label = f"{old_tdl} -> {new_tdl}"
        link_customdata.append(
            {
                "upgrade": f"{old_schema} -> {new_schema}",
                "transition": transition_label,
                "count": f"{count:,}" if count >= 1 else "",
            }
        )
        link_hovertemplates.append(
            "<extra></extra>"
            if is_anchor
            else "%{customdata.upgrade}<br>%{customdata.transition}<br>%{customdata.count} UniProt targets<extra></extra>"
        )

    annotations = [
        {
            "xref": "paper",
            "yref": "paper",
            "x": 0.02 + i * (0.96 / max(1, len(stages) - 1)),
            "y": 1.08,
            "text": stage,
            "showarrow": False,
            "xanchor": "center",
            "font": {"size": 12, "color": "#1f2933"},
        }
        for i, stage in enumerate(stages)
    ]

    data = [
        {
            "type": "sankey",
            "arrangement": "fixed",
            "node": {
                "pad": 20,
                "thickness": 16,
                "line": {"color": "rgba(31,41,51,0.35)", "width": 0.5},
                "label": labels,
                "color": colors,
                "x": xs,
                "y": ys,
                "customdata": customdata,
                "hovertemplate": "%{customdata}<br>Total flow through node: %{value:,}<extra></extra>",
            },
            "link": {
                "source": sources,
                "target": targets,
                "value": values,
                "color": link_colors,
                "customdata": link_customdata,
                "hovertemplate": link_hovertemplates,
            },
        }
    ]

    layout = {
        "title": {
            "text": "TDL Changes by UniProt Across TCRD / Pharos Versions",
            "x": 0.02,
            "xanchor": "left",
        },
        "font": {"family": "Arial, sans-serif", "size": 12, "color": "#1f2933"},
        "width": 1580,
        "height": 820,
        "paper_bgcolor": "white",
        "plot_bgcolor": "white",
        "margin": {"l": 30, "r": 30, "t": 85, "b": 30},
        "annotations": annotations
        + [
            {
                "xref": "paper",
                "yref": "paper",
                "x": 0.02,
                "y": 1.16,
            "text": "Bars show total UniProt targets per TDL. Tvoid captures targets newly annotated or absent/obsolete between versions.",
                "showarrow": False,
                "xanchor": "left",
                "font": {"size": 13, "color": "#52616b"},
            }
        ],
    }

    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <title>Pharos TDL Changes Sankey</title>
  <script src="https://cdn.plot.ly/plotly-2.35.2.min.js"></script>
  <style>
    body {{ margin: 0; background: #fff; font-family: Arial, sans-serif; }}
    #chart {{ width: 100vw; height: 100vh; min-height: 760px; }}
  </style>
</head>
<body>
  <div id="chart"></div>
  <script>
    const data = {json.dumps(data)};
    const layout = {json.dumps(layout)};
    Plotly.newPlot("chart", data, layout, {{responsive: true, displaylogo: false}});
  </script>
</body>
</html>
"""


def render_d3_html(stages, rows):
    nodes = []
    node_ids = set()
    stage_index = {stage: i for i, stage in enumerate(stages)}
    tdl_index = {tdl: i for i, tdl in enumerate(TDL_ORDER)}
    for stage in stages:
        for tdl in TDL_ORDER:
            node_id = f"{stage}|{tdl}"
            node_ids.add(node_id)
            nodes.append(
                {
                    "id": node_id,
                    "stage": stage,
                    "tdl": tdl,
                    "color": TDL_COLORS[tdl],
                    "stageIndex": stage_index[stage],
                    "tdlIndex": tdl_index[tdl],
                }
            )

    links = []
    for old_schema, new_schema, old_tdl, new_tdl, count in rows:
        source = f"{old_schema}|{old_tdl}"
        target = f"{new_schema}|{new_tdl}"
        if source not in node_ids or target not in node_ids:
            continue
        is_anchor = old_tdl == "Tvoid" and new_tdl == "Tvoid" and count < 1
        if old_tdl == "Tvoid" and new_tdl != "Tvoid":
            transition = f"Tvoid -> {new_tdl} newly annotated"
        elif new_tdl == "Tvoid" and old_tdl != "Tvoid":
            transition = f"{old_tdl} -> Tvoid removed/obsolete"
        elif old_tdl == new_tdl:
            transition = f"{old_tdl} unchanged"
        else:
            transition = f"{old_tdl} -> {new_tdl}"
        links.append(
            {
                "source": source,
                "target": target,
                "value": count,
                "color": "rgba(0,0,0,0)" if is_anchor else rgba(TDL_COLORS[old_tdl], 0.34),
                "upgrade": f"{old_schema} -> {new_schema}",
                "transition": transition,
                "countLabel": "" if is_anchor else f"{int(count):,}",
                "isAnchor": is_anchor,
            }
        )

    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <title>Pharos TDL Changes Sankey</title>
  <script src="https://cdn.jsdelivr.net/npm/d3@7"></script>
  <script src="https://cdn.jsdelivr.net/npm/d3-sankey@0.12.3/dist/d3-sankey.min.js"></script>
  <style>
    body {{ margin: 0; background: #fff; font-family: Arial, sans-serif; color: #1f2933; }}
    #wrap {{ width: calc(100vw - 36px); height: calc(100vh - 28px); min-height: 760px; padding: 14px 18px; box-sizing: border-box; }}
    .title {{ font-size: 22px; font-weight: 700; }}
    .subtitle {{ font-size: 13px; fill: #52616b; }}
    .stage {{ font-size: 12px; font-weight: 700; fill: #1f2933; }}
    .node-label {{ font-size: 11px; font-weight: 700; pointer-events: none; }}
    .tooltip {{
      position: fixed; pointer-events: none; opacity: 0; background: rgba(31,41,51,0.94);
      color: #fff; padding: 8px 10px; border-radius: 4px; font-size: 12px; line-height: 1.35;
      max-width: 280px;
    }}
  </style>
</head>
<body>
<div id="wrap"></div>
<div id="tip" class="tooltip"></div>
<script>
const stages = {json.dumps(stages)};
const tdlOrder = {json.dumps(TDL_ORDER)};
const data = {{
  nodes: {json.dumps(nodes)},
  links: {json.dumps(links)}
}};

function draw() {{
  const wrap = document.getElementById("wrap");
  wrap.innerHTML = "";
  const width = Math.max(1200, wrap.clientWidth);
  const height = Math.max(760, wrap.clientHeight);
  const margin = {{top: 88, right: 28, bottom: 24, left: 28}};
  const svg = d3.select(wrap).append("svg")
    .attr("width", width)
    .attr("height", height)
    .attr("viewBox", [0, 0, width, height]);

  svg.append("text").attr("x", margin.left).attr("y", 30).attr("class", "title")
    .text("TDL Changes by UniProt Across TCRD / Pharos Versions");
  svg.append("text").attr("x", margin.left).attr("y", 54).attr("class", "subtitle")
    .text("Bars show total UniProt targets per TDL. Tvoid captures targets newly annotated or absent/obsolete between versions.");

  const graph = {{
    nodes: data.nodes.map(d => ({{...d}})),
    links: data.links.map(d => ({{...d}}))
  }};

  const sankey = d3.sankey()
    .nodeId(d => d.id)
    .nodeWidth(16)
    .nodePadding(12)
    .nodeAlign((node, n) => node.stageIndex)
    .nodeSort((a, b) => a.tdlIndex - b.tdlIndex)
    .extent([[margin.left, margin.top], [width - margin.right, height - margin.bottom]]);

  sankey(graph);

  const tip = d3.select("#tip");
  const showTip = (event, html) => {{
    tip.html(html).style("opacity", 1);
    const el = tip.node();
    const pad = 14;
    const box = el.getBoundingClientRect();
    let left = event.clientX + 14;
    let top = event.clientY + 14;
    if (left + box.width + pad > window.innerWidth) left = event.clientX - box.width - 14;
    if (top + box.height + pad > window.innerHeight) top = event.clientY - box.height - 14;
    left = Math.max(pad, Math.min(left, window.innerWidth - box.width - pad));
    top = Math.max(pad, Math.min(top, window.innerHeight - box.height - pad));
    tip.style("left", `${{left}}px`).style("top", `${{top}}px`);
  }};
  const hideTip = () => tip.style("opacity", 0);

  svg.append("g")
    .attr("fill", "none")
    .selectAll("path")
    .data(graph.links.filter(d => !d.isAnchor))
    .join("path")
      .attr("d", d3.sankeyLinkHorizontal())
      .attr("stroke", d => d.color)
      .attr("stroke-width", d => Math.max(1, d.width))
      .attr("stroke-opacity", 1)
      .on("mousemove", (event, d) => showTip(event,
        `<strong>${{d.upgrade}}</strong><br>${{d.transition}}<br>${{d.countLabel}} UniProt targets`))
      .on("mouseleave", hideTip);

  const node = svg.append("g")
    .selectAll("g")
    .data(graph.nodes)
    .join("g");

  node.append("rect")
    .attr("x", d => d.x0)
    .attr("y", d => d.y0)
    .attr("height", d => Math.max(1, d.y1 - d.y0))
    .attr("width", d => d.x1 - d.x0)
    .attr("rx", 2)
    .attr("fill", d => d.color)
    .attr("stroke", "rgba(31,41,51,0.35)")
    .on("mousemove", (event, d) => showTip(event,
      `<strong>${{d.stage}} ${{d.tdl}}</strong><br>${{Math.round(d.value).toLocaleString()}} UniProt targets`))
    .on("mouseleave", hideTip);

  node.append("text")
    .attr("class", "node-label")
    .attr("x", d => d.x0 + 20)
    .attr("y", d => (d.y0 + d.y1) / 2 + 4)
    .text(d => d.tdl)
    .attr("fill", "#1f2933");

  const xByStage = new Map();
  for (const node of graph.nodes) {{
    if (!xByStage.has(node.stage)) xByStage.set(node.stage, (node.x0 + node.x1) / 2);
  }}
  svg.append("g")
    .selectAll("text")
    .data(stages)
    .join("text")
      .attr("class", "stage")
      .attr("x", d => xByStage.get(d))
      .attr("y", margin.top - 18)
      .attr("text-anchor", "middle")
      .text(d => d);
}}

draw();
window.addEventListener("resize", draw);
</script>
</body>
</html>
"""


def main():
    stages, rows = parse_summary()
    svg = render_svg(stages, rows)
    SVG_OUT.write_text(svg)
    HTML_OUT.write_text(
        "<!doctype html><meta charset='utf-8'><title>Pharos TDL Sankey</title>"
        "<body style='margin:0;background:#fff'>"
        + svg
        + "</body>"
    )
    full_stages, full_rows = parse_full_flows()
    PLOTLY_OUT.write_text(render_d3_html(full_stages, full_rows))
    print(SVG_OUT.resolve())
    print(HTML_OUT.resolve())
    print(PLOTLY_OUT.resolve())


if __name__ == "__main__":
    main()
